import os
import sys
import pandas as pd
import argparse
import glob
import ast
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# --- Constants ---

MODEL_CONFIG = {
    "logistic_regression": {"name": "Logistic Regression", "interpretable": "Yes"},
    "svm": {"name": "Support Vector Machine", "interpretable": "Yes"},
    "random_forest": {"name": "Random Forest", "interpretable": "No"},
    "decision_tree": {"name": "Decision Tree", "interpretable": "Yes"},
    "knn": {"name": "K-Nearest-Neighbors", "interpretable": "No"},
}

PREDICTOR_ALIASES = {
    "gad7_sum": "GAD-7",
    "sds_sum": "SDS",
    "base_phq9": "PHQ-9",
}

COLUMN_ORDER = [
    "Model type",
    "Interp?",
    "AUC",
    "Accuracy",
    "Depth",
    "Pred.",
    "Coefficient",
]

# --- Helper Functions ---

def get_args(argv):
    """Parses command-line arguments."""
    parser = argparse.ArgumentParser(
        description="Generate a summary table from machine learning model output files."
    )
    parser.add_argument(
        "--input_dir",
        type=str,
        required=True,
        help="Path to the input directory containing the results.",
    )
    parser.add_argument(
        "--subdirectory",
        type=str,
        default="results_Brighten-v1",
        help="Subdirectory within the input directory containing the .csv files.",
    )
    parser.add_argument(
        "--output_filename",
        type=str,
        default=None,
        help="Base name for the output file (without extension). If not provided, it will be derived from the input directory name.",
    )
    parser.add_argument(
        "--sort_sections",
        action="store_true",
        help="Sort the model sections based on the highest AUC in each section.",
    )
    parser.add_argument(
        "--no_excel",
        action="store_true",
        help="Force CSV output even if openpyxl is installed.",
    )
    parser.add_argument(
        "--decimal_places",
        type=int,
        default=2,
        help="Number of decimal places to round to.",
    )
    return parser.parse_args(argv)

def load_and_validate_data(source_dir):
    """Loads and concatenates all CSV files from the source directory."""
    csv_files = glob.glob(os.path.join(source_dir, "*.csv"))
    if not csv_files:
        raise FileNotFoundError(f"No CSV files found in directory: {source_dir}")

    all_dfs = []
    required_cols = {"model", "auc", "ci_lower", "ci_upper", "predictors", "complexity", "mean_acc", "acc_ci_lower", "acc_ci_upper"}
    
    for f in csv_files:
        try:
            df = pd.read_csv(f)
            if not required_cols.issubset(df.columns):
                missing = required_cols - set(df.columns)
                raise ValueError(f"File '{f}' is missing required columns: {missing}")
            all_dfs.append(df)
        except Exception as e:
            print(f"Warning: Could not process file {f}. Skipping. Reason: {e}")

    if not all_dfs:
        raise ValueError("No valid data could be loaded from any CSV file.")

    return pd.concat(all_dfs, ignore_index=True)

def format_row(model_result, model_key, decimal_places):
    """Formats a single model result row for the output table."""
    try:
        # Safely evaluate the string representation of the list
        predictors_list = ast.literal_eval(model_result["predictors"])
        
        aliased_predictors = []
        for p in predictors_list:
            if p not in PREDICTOR_ALIASES:
                print(f"Warning: No alias in dict for {p}")
            aliased_predictors.append(PREDICTOR_ALIASES.get(p, p))

        predictors_str = ", ".join(aliased_predictors)
    except (ValueError, SyntaxError):
        predictors_str = model_result["predictors"] # Keep as is if parsing fails

    auc = model_result["auc"]
    ci_lower = model_result["ci_lower"]
    ci_upper = model_result["ci_upper"]

    acc = model_result["mean_acc"]
    acc_ci_lower = model_result["acc_ci_lower"]
    acc_ci_upper = model_result["acc_ci_upper"]

    return {
        "Model type": MODEL_CONFIG[model_key]["name"],
        "Interp?": MODEL_CONFIG[model_key]["interpretable"],
        "AUC": f"{auc:.{decimal_places}f} ({ci_lower + 0.5:.{decimal_places}f}, {ci_upper + 0.5:.{decimal_places}f})",
        "Accuracy": f"{acc:.{decimal_places}f} ({acc_ci_lower:.{decimal_places}f}, {acc_ci_upper:.{decimal_places}f})",
        "Depth": model_result["complexity"] if model_key in ["decision_tree", "random_forest"] else "-",
        "Pred.": predictors_str,
        "Coefficient": "TODO" if model_key in ["logistic_regression", "svm"] else "-",
        "_model_key": model_key, # Internal key for grouping
        "_auc_raw": auc, # Internal value for sorting
    }

def create_placeholder_row(model_key):
    """Creates a placeholder row for a model type with no significant results."""
    return {
        "Model type": MODEL_CONFIG[model_key]["name"],
        "Interp?": MODEL_CONFIG[model_key]["interpretable"],
        "AUC": "N.S.",
        "Accuracy": "N.S.",
        "Depth": "-",
        "Pred.": "N.S.",
        "Coefficient": "N.S." if model_key in ["logistic_regression", "svm"] else "-",
        "_model_key": model_key,
        "_auc_raw": -1.0, # Ensure it sorts last
    }

def write_to_excel(df, output_path):
    """Writes the DataFrame to a formatted Excel file."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Border, Side, PatternFill
        from openpyxl.utils.dataframe import dataframe_to_rows
    except ImportError:
        print("Warning: `openpyxl` is not installed. Cannot create .xlsx file.")
        raise

    wb = Workbook()
    ws = wb.active
    ws.title = "Model Summary"

    # --- Styles ---
    header_font = Font(bold=True)
    grey_font = Font(color="9b9b9b")
    thin_border_bottom = Border(bottom=Side(style='thin'))

    # --- Write Data ---
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        ws.append(row)
        # Style header
        if r_idx == 1:
            for cell in ws[r_idx]:
                cell.font = header_font
                cell.border = thin_border_bottom
            continue

        # --- Apply Formatting ---
        current_model_type = ws.cell(row=r_idx, column=1).value
        is_first_in_section = current_model_type not in [None, ""]

        # Add border to the top of a new section (on the first row of it)
        if is_first_in_section and r_idx > 2:
            for cell in ws[r_idx]:
                if cell.border.top is None or cell.border.top.style is None:
                    cell.border = Border(top=Side(style='thin'))

        # Grey out text for non-primary models in a section
        if not is_first_in_section:
            for cell in ws[r_idx]:
                cell.font = grey_font

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(output_path)


# --- Main Execution ---

def main():
    """Main script execution."""
    args = get_args(sys.argv[1:])

    # --- Path Validation ---
    input_dir = args.input_dir
    source_dir = os.path.join(input_dir, args.subdirectory)
    if not os.path.isdir(input_dir):
        raise FileNotFoundError(f"Input directory not found: {input_dir}")
    if not os.path.isdir(source_dir):
        raise FileNotFoundError(f"Subdirectory not found: {source_dir}")

    # --- Data Loading ---
    try:
        master_df = load_and_validate_data(source_dir)
    except (FileNotFoundError, ValueError) as e:
        print(f"Error: {e}")
        sys.exit(1)

    # --- Data Processing ---
    model_sections = []
    for key, config in MODEL_CONFIG.items():
        model_df = master_df[master_df["model"] == key]
        
        section_rows = []
        if not model_df.empty:
            # Sort models within the section by AUC
            model_df_sorted = model_df.sort_values("auc", ascending=False)
            for _, row_data in model_df_sorted.iterrows():
                section_rows.append(format_row(row_data, key, args.decimal_places))
        else:
            # Add placeholder if no models of this type were found
            section_rows.append(create_placeholder_row(key))
        
        model_sections.append(section_rows)

    # --- Section Sorting ---
    if args.sort_sections:
        # Sort by the _auc_raw of the first (best) model in each section
        model_sections.sort(key=lambda s: s[0]["_auc_raw"], reverse=True)

    # --- Final Table Assembly ---
    final_rows = []
    for section in model_sections:
        for i, row in enumerate(section):
            if i > 0: # For display purposes, clear redundant info
                row["Model Type"] = ""
                row["Interp?"] = ""
            final_rows.append(row)
    
    final_df = pd.DataFrame(final_rows)
    # Drop internal columns and reorder
    final_df = final_df.drop(columns=["_model_key", "_auc_raw"])[COLUMN_ORDER]
    
    # --- Output Generation ---
    if args.output_filename is None:
        leaf_dir = os.path.basename(os.path.normpath(args.input_dir))
        output_filename = f"{leaf_dir}_summary_table"
    else:
        output_filename = args.output_filename

    output_basename = os.path.join(input_dir, output_filename)
    use_excel = not args.no_excel
    
    if use_excel:
        try:
            write_to_excel(final_df, f"{output_basename}.xlsx")
            print(f"Successfully created formatted Excel file: {output_basename}.xlsx")
        except (ImportError, Exception) as e:
            print(f"Could not write to Excel ({e}). Falling back to CSV.")
            final_df.to_csv(f"{output_basename}.csv", index=False)
            print(f"Successfully created CSV file: {output_basename}.csv")
    else:
        final_df.to_csv(f"{output_basename}.csv", index=False)
        print(f"Successfully created CSV file: {output_basename}.csv")

    print(f"\nGenerated table with {len(final_df)} rows and {len(final_df.columns)} columns.")

if __name__ == "__main__":
    main() 